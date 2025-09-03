import json
import os
from django.core.paginator import Paginator, EmptyPage
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from .models import MainCategory, SubCategory, Article, Tag, Library, Module, OperationType, Function, Parameter

from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages

# ========== 页面渲染视图 ==========

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('Pythonfun:category_management') # 使用URL名称进行重定向
        else:
            messages.error(request, '用户名或密码无效')
            return render(request, 'admin/登录.html', status=401)
    return render(request, 'admin/登录.html')

def index_view(request):
    """语法页面视图 - 显示语法类型的文章"""
    # 获取语法类型的文章
    articles = Article.objects.filter(
        content_type=Article.ContentType.GRAMMAR,
        is_published=True
    ).order_by('-created_at')
    
    # 获取分类树
    main_categories = MainCategory.objects.filter(is_enabled=True).order_by('order')
    category_tree = []
    for main_category in main_categories:
        # 获取该主分类下有语法类型文章的子分类
        sub_categories_with_articles = SubCategory.objects.filter(
            parent=main_category,
            is_enabled=True,
            article__content_type=Article.ContentType.GRAMMAR,
            article__is_published=True
        ).distinct().order_by('id')
        
        # 只有当主分类下有文章时，才添加到分类树中
        if sub_categories_with_articles.exists():
            sub_categories_with_count = []
            for sub in sub_categories_with_articles:
                article_count = Article.objects.filter(
                    category=sub,
                    content_type=Article.ContentType.GRAMMAR,
                    is_published=True
                ).count()
                sub.article_count = article_count
                sub_categories_with_count.append(sub)
            
            category_tree.append({
                'main_category': main_category,
                'sub_categories': sub_categories_with_count
            })
    
    # 获取当前文章（用于显示默认文章）
    current_article = Article.objects.filter(
        content_type=Article.ContentType.GRAMMAR,
        is_published=True
    ).order_by('created_at').first()
    
    # 如果没有已发布的文章，且用户是管理员，尝试获取未发布的文章（用于管理员预览）
    if not current_article and request.user.is_authenticated and request.user.is_staff:
        current_article = Article.objects.filter(
            content_type=Article.ContentType.GRAMMAR,
            is_published=False
        ).order_by('created_at').first()
    
    context = {
        'articles': articles,
        'category_tree': category_tree,
        'current_article': current_article,
        'content_type': 'grammar'
    }
    return render(request, 'front/index.html', context)

def category_view(request, slug):
    """子分类文章显示视图 - 根据内容类型显示到不同页面"""
    try:
        content_type = request.GET.get('type', 'grammar')
        content_type_map = {
            'grammar': Article.ContentType.GRAMMAR,
            'data-structure': Article.ContentType.DATA_STRUCTURE,
            'ai-programming': Article.ContentType.AI_PROGRAMMING
        }
        model_content_type = content_type_map.get(content_type, Article.ContentType.GRAMMAR)
        
        # 获取当前分类
        try:
            current_category = SubCategory.objects.get(slug=slug, is_enabled=True)
        except SubCategory.DoesNotExist:
            return render(request, 'front/404.html', {
                'error_message': f'分类 "{slug}" 不存在或已被禁用'
            }, status=404)
        
        # 获取当前分类的文章
        current_article = Article.objects.filter(
            category=current_category,
            content_type=model_content_type,
            is_published=True
        ).order_by('created_at').first()
        
        # 如果没有已发布的文章，且用户是管理员，尝试获取未发布的文章（用于管理员预览）
        if not current_article and request.user.is_authenticated and request.user.is_staff:
            current_article = Article.objects.filter(
                category=current_category,
                content_type=model_content_type,
                is_published=False
            ).order_by('created_at').first()
        
        # 构建分类树，只包含有指定内容类型文章的分类
        category_tree = []
        
        # 获取包含指定内容类型文章的子分类
        sub_categories_with_articles = SubCategory.objects.filter(
            article__content_type=model_content_type,
            article__is_published=True,
            is_enabled=True
        ).distinct().order_by('parent__id', 'id')
        
        # 按主分类分组
        current_main_category = None
        current_sub_categories = []
        
        for sub_category in sub_categories_with_articles:
            if current_main_category != sub_category.parent:
                # 保存前一个主分类的数据
                if current_main_category and current_sub_categories:
                    category_tree.append({
                        'main_category': current_main_category,
                        'sub_categories': current_sub_categories
                    })
                
                # 开始新的主分类
                current_main_category = sub_category.parent
                current_sub_categories = []
            
            # 计算文章数量
            article_count = Article.objects.filter(
                category=sub_category,
                content_type=model_content_type,
                is_published=True
            ).count()
            sub_category.article_count = article_count
            current_sub_categories.append(sub_category)
        
        # 添加最后一个主分类的数据
        if current_main_category and current_sub_categories:
            category_tree.append({
                'main_category': current_main_category,
                'sub_categories': current_sub_categories
            })
        
        # 根据内容类型选择模板
        template_map = {
            'grammar': 'front/index.html',
            'data-structure': 'front/数据结构.html',
            'ai-programming': 'front/AI编程.html'
        }
        
        template_name = template_map.get(content_type, 'front/index.html')
        
        # 构建上下文
        context = {
            'category_tree': category_tree,
            'current_category': current_category,
            'current_article': current_article,
            'content_type': content_type
        }
        
        # 如果没有文章，添加提示信息
        if not current_article:
            context['no_articles_message'] = f'分类 "{current_category.name}" 暂无{content_type}类型的文章'
        else:
            # 如果有文章，确保文章内容被正确处理
            if hasattr(current_article, 'content') and current_article.content:
                # 如果文章有content字段，转换为HTML
                import markdown
                current_article.content_html = markdown.markdown(current_article.content)
        
        return render(request, template_name, context)
        
    except Exception as e:
        # 记录错误并返回友好的错误页面
        print(f"Category view error: {str(e)}")
        return render(request, 'front/404.html', {
            'error_message': '加载分类页面时发生错误，请稍后重试'
        }, status=500)

@login_required
def category_management_view(request):
    return render(request, 'admin/分类管理.html')

@login_required
def article_edit_view(request):
    article_id = request.GET.get('id')
    article = None
    if article_id:
        article = get_object_or_404(Article, id=article_id)
    return render(request, 'admin/文章编辑.html', {'article': article})

@login_required
def course_management_view(request):
    return render(request, 'admin/教程管理.html')

def tutorial_detail_view(request, pk):
    tutorial = get_object_or_404(Article, pk=pk)
    is_admin = request.user.is_authenticated and request.user.is_staff
    if not tutorial.is_published and not is_admin:
        return render(request, 'front/404.html', {'error_message': '教程不存在或未发布'})

    related_tutorials = Article.objects.filter(
        category=tutorial.category, content_type=Article.ContentType.GRAMMAR, is_published=True
    ).exclude(pk=pk)[:5]
    
    context = {
        'tutorial': tutorial,
        'related_tutorials': related_tutorials,
        'is_admin': is_admin
    }
    return render(request, 'front/tutorial_detail.html', context)

# ========== API 视图 ==========

@require_http_methods(["GET", "POST"])
def main_category_api(request):
    if request.method == 'GET':
        page = request.GET.get('page', 1)
        page_size = 10
        
        categories = MainCategory.objects.all().order_by('order')
        
        # 添加分页
        paginator = Paginator(categories, page_size)
        try:
            categories_page = paginator.page(page)
        except EmptyPage:
            categories_page = paginator.page(paginator.num_pages)
        
        data = {
            'items': list(categories_page.object_list.values('id', 'name', 'slug', 'order', 'is_enabled')),
            'current_page': categories_page.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count
        }
        return JsonResponse(data, safe=False)
    elif request.method == 'POST':
        data = json.loads(request.body)
        try:
            main_category = MainCategory.objects.create(
                name=data['name'],
                slug=data['slug'],
                order=data.get('order', 0),
                is_enabled=data.get('is_enabled', True)
            )
            return JsonResponse({
                'status': 'success', 
                'message': 'Main category created successfully', 
                'id': main_category.id
            }, status=201)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)

@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def main_category_detail_api(request, pk):
    main_category = get_object_or_404(MainCategory, pk=pk)
    if request.method == 'GET':
        data = {
            'id': main_category.id,
            'name': main_category.name,
            'slug': main_category.slug,
            'order': main_category.order,
            'is_enabled': main_category.is_enabled
        }
        return JsonResponse(data)
    elif request.method == 'PUT':
        data = json.loads(request.body)
        main_category.name = data.get('name', main_category.name)
        main_category.slug = data.get('slug', main_category.slug)
        main_category.order = data.get('order', main_category.order)
        main_category.is_enabled = data.get('is_enabled', main_category.is_enabled)
        main_category.save()
        return JsonResponse({'status': 'success', 'message': 'Main category updated successfully'})
    elif request.method == 'DELETE':
        # 检查是否有子分类
        if main_category.subcategories.exists():
            return JsonResponse({
                'status': 'error', 
                'message': '无法删除主分类：该主分类下还有子分类，请先删除所有子分类'
            }, status=400)
        
        # 检查是否有相关文章
        if Article.objects.filter(category__parent=main_category).exists():
            return JsonResponse({
                'status': 'error', 
                'message': '无法删除主分类：该主分类下还有文章，请先删除所有文章'
            }, status=400)
        
        main_category.delete()
        return JsonResponse({'status': 'success', 'message': 'Main category deleted successfully'})

@csrf_exempt
@require_http_methods(["GET", "POST"])
def sub_category_api(request):
    if request.method == 'GET':
        main_category_id = request.GET.get('main_category_id')
        page = request.GET.get('page', 1)
        page_size = 10 # Assuming a default page size

        categories = SubCategory.objects.all()
        if main_category_id:
            categories = categories.filter(parent_id=main_category_id)

        # Add pagination
        paginator = Paginator(categories, page_size)
        try:
            categories_page = paginator.page(page)
        except EmptyPage:
            categories_page = paginator.page(paginator.num_pages)

        data = {
            'items': [{
                'id': cat.id,
                'name': cat.name,
                'slug': cat.slug,
                'parent_id': cat.parent.id if cat.parent else None,
                'parent_name': cat.parent.name if cat.parent else None,
                'is_enabled': cat.is_enabled
            } for cat in categories_page.object_list],
            'current_page': categories_page.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count
        }
        return JsonResponse(data, safe=False)
    elif request.method == 'POST':
        data = json.loads(request.body)
        parent_id = data.get('parent_id')
        parent_category = get_object_or_404(MainCategory, pk=parent_id) if parent_id else None
        sub_category = SubCategory.objects.create(
            name=data['name'],
            slug=data['slug'],
            parent=parent_category,
            is_enabled=data.get('is_enabled', True)
        )
        return JsonResponse({'status': 'success', 'message': 'Sub category created successfully', 'id': sub_category.id}, status=201)

@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def sub_category_detail_api(request, pk):
    sub_category = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'GET':
        data = {
            'id': sub_category.id,
            'name': sub_category.name,
            'slug': sub_category.slug,
            'parent_id': sub_category.parent.id if sub_category.parent else None,
            'is_enabled': sub_category.is_enabled
        }
        return JsonResponse(data)
    elif request.method == 'PUT':
        data = json.loads(request.body)
        sub_category.name = data.get('name', sub_category.name)
        sub_category.slug = data.get('slug', sub_category.slug)  # 添加slug字段更新
        if 'parent_id' in data:
            parent_id = data.get('parent_id')
            if parent_id:
                sub_category.parent = get_object_or_404(MainCategory, pk=parent_id)
            else:
                sub_category.parent = None
        sub_category.is_enabled = data.get('is_enabled', sub_category.is_enabled)
        sub_category.save()
        return JsonResponse({'status': 'success', 'message': 'Sub category updated successfully'})
    elif request.method == 'DELETE':
        # 检查是否有相关文章
        if Article.objects.filter(category=sub_category).exists():
            return JsonResponse({
                'status': 'error', 
                'message': '无法删除子分类：该子分类下还有文章，请先删除所有文章'
            }, status=400)
        
        sub_category.delete()
        return JsonResponse({'status': 'success', 'message': 'Sub category deleted successfully'})

@csrf_exempt
@require_http_methods(["GET", "POST"])
def article_api(request):
    if request.method == 'GET':
        articles = Article.objects.all()
        data = [{
            'id': article.id,
            'title': article.title,
            'subtitle': article.subtitle,
            'summary': article.summary,
            'read_time_minutes': article.read_time_minutes,
            'is_published': article.is_published,
            'created_at': article.created_at.isoformat(),
            'updated_at': article.updated_at.isoformat(),
            'category_name': article.category.name if article.category else None,
            'tags': [tag.name for tag in article.tags.all()]
        } for article in articles]
        return JsonResponse(data, safe=False)
    elif request.method == 'POST':
        try:
            print("收到POST请求到article_api")
            data = json.loads(request.body)
            print("解析的请求数据:", data)
            
            # 验证必填字段
            errors = {}
            if not data.get('title', '').strip():
                errors['title'] = ['文章标题不能为空']
            if not data.get('summary', '').strip():
                errors['summary'] = ['文章摘要不能为空']
            if not data.get('category_id'):
                errors['category'] = ['请选择文章分类']
            
            # 验证内容
            content_html = data.get('content_html', '').strip()
            content_code = data.get('content_code', '').strip()
            if not content_html and not content_code:
                errors['content'] = ['请至少添加一些内容（富文本或代码）']
            
            # 如果有验证错误，返回错误信息
            if errors:
                return JsonResponse({
                    'status': 'error',
                    'message': '验证失败',
                    'errors': errors
                }, status=400)
            
            # 创建文章
            article = Article.objects.create(
                title=data['title'].strip(),
                subtitle=data.get('subtitle', '').strip(),
                summary=data.get('summary', '').strip(),
                content_html=content_html,
                content_code=content_code,
                code_language=data.get('code_language', 'python'),
                content_type=data.get('content_type', Article.ContentType.GRAMMAR),
                read_time_minutes=data.get('read_time_minutes', 5),
                is_published=data.get('is_published', False)
            )
            
            # 设置分类
            if data.get('category_id'):
                try:
                    article.category = get_object_or_404(SubCategory, pk=data.get('category_id'))
                    article.save()
                except Exception as e:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'选择的分类不存在: {str(e)}'
                    }, status=400)
            
            # 设置标签
            if data.get('tags'):
                article.tags.set(Tag.objects.filter(name__in=data.get('tags')))
            
            return JsonResponse({
                'status': 'success', 
                'message': '文章创建成功', 
                'id': article.id
            }, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': '请求数据格式错误'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'保存失败: {str(e)}'
            }, status=500)

@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def article_detail_api(request, pk):
    article = get_object_or_404(Article, pk=pk)
    if request.method == 'GET':
        data = {
            'id': article.id,
            'title': article.title,
            'subtitle': article.subtitle,
            'summary': article.summary,
            'read_time_minutes': article.read_time_minutes,
            'content_html': article.content_html,
            'content_code': article.content_code,
            'code_language': article.code_language,
            'category_id': article.category.id if article.category else None,
            'tags': [tag.name for tag in article.tags.all()],
        }
        return JsonResponse(data)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            
            # 验证必填字段
            errors = {}
            if not data.get('title', '').strip():
                errors['title'] = ['文章标题不能为空']
            if not data.get('summary', '').strip():
                errors['summary'] = ['文章摘要不能为空']
            if not data.get('category_id'):
                errors['category'] = ['请选择文章分类']
            
            # 验证内容
            content_html = data.get('content_html', '').strip()
            content_code = data.get('content_code', '').strip()
            if not content_html and not content_code:
                errors['content'] = ['请至少添加一些内容（富文本或代码）']
            
            # 如果有验证错误，返回错误信息
            if errors:
                return JsonResponse({
                    'status': 'error',
                    'message': '验证失败',
                    'errors': errors
                }, status=400)
            
            # 更新文章
            article.title = data.get('title', article.title).strip()
            article.subtitle = data.get('subtitle', article.subtitle).strip()
            article.summary = data.get('summary', article.summary).strip()
            article.read_time_minutes = data.get('read_time_minutes', article.read_time_minutes)
            article.content_html = content_html
            article.content_code = content_code
            article.code_language = data.get('code_language', article.code_language)
            
            # 设置分类
            if data.get('category_id'):
                try:
                    article.category = get_object_or_404(SubCategory, pk=data.get('category_id'))
                except Exception as e:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'选择的分类不存在: {str(e)}'
                    }, status=400)
            
            article.save()
            
            # 设置标签
            if data.get('tags'):
                article.tags.set(Tag.objects.filter(name__in=data.get('tags', [])))
            
            return JsonResponse({
                'status': 'success', 
                'message': '文章更新成功'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': '请求数据格式错误'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'更新失败: {str(e)}'
            }, status=500)

@csrf_exempt
@require_http_methods(["GET", "POST"])
def course_api(request):
    if request.method == 'GET':
        page = request.GET.get('page', 1)
        page_size = 10 # Assuming a default page size

        # 显示所有文章，不再限制内容类型
        courses = Article.objects.all()

        # Add pagination
        paginator = Paginator(courses, page_size)
        try:
            courses_page = paginator.page(page)
        except EmptyPage:
            courses_page = paginator.page(paginator.num_pages)

        data = {
            'items': [{
                'id': course.id,
                'title': course.title,
                'subtitle': course.subtitle,
                'summary': course.summary,
                'read_time_minutes': course.read_time_minutes,
                'is_published': course.is_published,
                'created_at': course.created_at.isoformat(),
                'updated_at': course.updated_at.isoformat(),
                'category_name': course.category.name if course.category else None,
                'content_type': course.content_type,
                'content_type_display': course.get_content_type_display(),
                'tags': [tag.name for tag in course.tags.all()]
            } for course in courses_page.object_list],
            'current_page': courses_page.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count
        }
        return JsonResponse(data, safe=False)
    elif request.method == 'POST':
        data = json.loads(request.body)
        course = Article.objects.create(
            title=data['title'],
            subtitle=data.get('subtitle', ''),
            summary=data.get('summary', ''),
            read_time_minutes=data.get('read_time_minutes', 0),
            content_html=data.get('content_html', ''),
            content_code=data.get('content_code', ''),
            code_language=data.get('code_language', ''),
            content_type=Article.ContentType.GRAMMAR,
            is_published=data.get('is_published', False)
        )
        if data.get('category_id'):
            course.category = get_object_or_404(SubCategory, pk=data.get('category_id'))
        course.save()
        if data.get('tags'):
            course.tags.set(Tag.objects.filter(name__in=data.get('tags')))
        return JsonResponse({'status': 'success', 'message': 'Course created successfully', 'id': course.id}, status=201)

@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def course_detail_api(request, pk):
    # 允许编辑所有类型的文章，不再限制内容类型
    course = get_object_or_404(Article, pk=pk)
    if request.method == 'GET':
        data = {
            'id': course.id,
            'title': course.title,
            'subtitle': course.subtitle,
            'summary': course.summary,
            'read_time_minutes': course.read_time_minutes,
            'content_html': course.content_html,
            'content_code': course.content_code,
            'code_language': course.code_language,
            'category_id': course.category.id if course.category else None,
            'category_name': course.category.name if course.category else None,
            'tags': [tag.name for tag in course.tags.all()],
            'is_published': course.is_published,
        }
        return JsonResponse(data)
    elif request.method == 'PUT':
        data = json.loads(request.body)
        course.title = data.get('title', course.title)
        course.subtitle = data.get('subtitle', course.subtitle)
        course.summary = data.get('summary', course.summary)
        course.read_time_minutes = data.get('read_time_minutes', course.read_time_minutes)

        course.content_html = data.get('content_html', course.content_html)
        course.content_code = data.get('content_code', course.content_code)
        course.code_language = data.get('code_language', course.code_language)
        if data.get('category_id'):
            course.category = get_object_or_404(SubCategory, pk=data.get('category_id'))
        course.save()
        course.tags.set(Tag.objects.filter(name__in=data.get('tags', [])))
        return JsonResponse({'status': 'success', 'message': 'Course updated successfully'})
    elif request.method == 'DELETE':
        try:
            # 记录删除的教程信息用于日志
            course_title = course.title
            course_id = course.id
            
            # 删除教程
            course.delete()
            
            return JsonResponse({
                'status': 'success', 
                'message': f'教程 "{course_title}" 已成功删除',
                'deleted_id': course_id
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'删除教程失败: {str(e)}'
            }, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def course_publish_api(request, pk):
    """发布/取消发布教程API"""
    try:
        # 允许发布/取消发布所有类型的文章，不再限制内容类型
        course = get_object_or_404(Article, pk=pk)
        data = json.loads(request.body)
        
        # 支持多种请求格式
        action = data.get('action')
        is_published = data.get('is_published')
        
        if action == 'publish':
            course.is_published = True
            course.save()
            return JsonResponse({
                'status': 'success', 
                'message': f'教程 "{course.title}" 发布成功！'
            })
        elif action == 'unpublish':
            course.is_published = False
            course.save()
            return JsonResponse({
                'status': 'success', 
                'message': f'教程 "{course.title}" 已取消发布'
            })
        elif is_published is not None:
            course.is_published = is_published
            course.save()
            return JsonResponse({
                'status': 'success', 
                'message': 'Course publish status updated successfully'
            })
        else:
            return JsonResponse({
                'status': 'error', 
                'message': 'Invalid request: missing action or is_published field'
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'status': 'error', 
            'message': f'发布失败: {str(e)}'
        }, status=500)

# ========== 导航栏页面视图 ==========

def function_library_view(request):
    """函数库页面视图"""
    return render(request, 'front/函数库.html')

def function_query_view(request):
    """函数查询页面视图"""
    return render(request, 'front/函数查询.html')

def data_structure_view(request):
    """数据结构页面视图"""
    # 获取数据结构类型的文章
    articles = Article.objects.filter(
        content_type=Article.ContentType.DATA_STRUCTURE,
        is_published=True
    ).order_by('-created_at')
    
    # 获取分类树
    main_categories = MainCategory.objects.filter(is_enabled=True).order_by('order')
    category_tree = []
    for main_category in main_categories:
        # 获取该主分类下有数据结构类型文章的子分类
        sub_categories_with_articles = SubCategory.objects.filter(
            parent=main_category,
            is_enabled=True,
            article__content_type=Article.ContentType.DATA_STRUCTURE,
            article__is_published=True
        ).distinct().order_by('id')
        
        # 只有当主分类下有文章时，才添加到分类树中
        if sub_categories_with_articles.exists():
            sub_categories_with_count = []
            for sub in sub_categories_with_articles:
                article_count = Article.objects.filter(
                    category=sub,
                    content_type=Article.ContentType.DATA_STRUCTURE,
                    is_published=True
                ).count()
                sub.article_count = article_count
                sub_categories_with_count.append(sub)
            
            category_tree.append({
                'main_category': main_category,
                'sub_categories': sub_categories_with_count
            })
    
    # 获取当前文章（用于显示默认文章）
    current_article = Article.objects.filter(
        content_type=Article.ContentType.DATA_STRUCTURE,
        is_published=True
    ).order_by('created_at').first()
    
    # 如果没有已发布的文章，且用户是管理员，尝试获取未发布的文章（用于管理员预览）
    if not current_article and request.user.is_authenticated and request.user.is_staff:
        current_article = Article.objects.filter(
            content_type=Article.ContentType.DATA_STRUCTURE,
            is_published=False
        ).order_by('created_at').first()
    
    context = {
        'articles': articles,
        'category_tree': category_tree,
        'current_article': current_article,
        'content_type': 'data_structure'
    }
    return render(request, 'front/数据结构.html', context)

def ai_programming_view(request):
    """AI编程页面视图"""
    # 获取AI编程类型的文章
    articles = Article.objects.filter(
        content_type=Article.ContentType.AI_PROGRAMMING,
        is_published=True
    ).order_by('-created_at')
    
    # 获取分类树
    main_categories = MainCategory.objects.filter(is_enabled=True).order_by('order')
    category_tree = []
    for main_category in main_categories:
        # 获取该主分类下有AI编程类型文章的子分类
        sub_categories_with_articles = SubCategory.objects.filter(
            parent=main_category,
            is_enabled=True,
            article__content_type=Article.ContentType.AI_PROGRAMMING,
            article__is_published=True
        ).distinct().order_by('id')
        
        # 只有当主分类下有文章时，才添加到分类树中
        if sub_categories_with_articles.exists():
            sub_categories_with_count = []
            for sub in sub_categories_with_articles:
                article_count = Article.objects.filter(
                    category=sub,
                    content_type=Article.ContentType.AI_PROGRAMMING,
                    is_published=True
                ).count()
                sub.article_count = article_count
                sub_categories_with_count.append(sub)
            
            category_tree.append({
                'main_category': main_category,
                'sub_categories': sub_categories_with_count
            })
    
    # 获取当前文章（用于显示默认文章）
    current_article = Article.objects.filter(
        content_type=Article.ContentType.AI_PROGRAMMING,
        is_published=True
    ).order_by('created_at').first()
    
    # 如果没有已发布的文章，且用户是管理员，尝试获取未发布的文章（用于管理员预览）
    if not current_article and request.user.is_authenticated and request.user.is_staff:
        current_article = Article.objects.filter(
            content_type=Article.ContentType.AI_PROGRAMMING,
            is_published=False
        ).order_by('created_at').first()
    
    context = {
        'articles': articles,
        'category_tree': category_tree,
        'current_article': current_article,
        'content_type': 'ai_programming'
    }
    return render(request, 'front/AI编程.html', context)

def project_view(request):
    """项目页面视图"""
    return render(request, 'front/项目.html')

def create_superuser_view(request):
    """创建超级用户页面视图"""
    return render(request, 'admin/create_superuser.html')

def reset_password_view(request):
    """重置密码页面视图"""
    return render(request, 'admin/reset_password.html')

@csrf_exempt
@require_http_methods(["POST"])
def create_superuser_api(request):
    """创建超级用户的API端点"""
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # 检查是否已存在超级用户
        if User.objects.filter(is_superuser=True).exists():
            return JsonResponse({
                'success': False,
                'message': '超级用户已存在'
            })
        
        # 创建超级用户
        username = "admin"
        email = "admin@example.com"
        password = "admin123456"
        
        # 检查用户是否已存在
        if User.objects.filter(username=username).exists():
            user = User.objects.get(username=username)
            user.is_superuser = True
            user.is_staff = True
            user.save()
            message = f"用户 {username} 已升级为超级用户"
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                is_superuser=True,
                is_staff=True
            )
            message = f"超级用户创建成功！用户名: {username}, 密码: {password}"
        
        return JsonResponse({
            'success': True,
            'message': message,
            'username': username,
            'password': password
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'创建超级用户失败: {str(e)}'
        })

@csrf_exempt
@require_http_methods(["POST"])
def reset_admin_password_api(request):
    """重置管理员密码的API端点"""
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # 查找admin用户
        try:
            user = User.objects.get(username='admin')
        except User.DoesNotExist:
            # 如果admin用户不存在，创建一个
            user = User.objects.create_user(
                username='admin',
                email='admin@example.com',
                password='admin123456',
                is_superuser=True,
                is_staff=True,
                is_active=True
            )
            message = "admin用户创建成功"
        else:
            # 重置密码
            user.set_password('admin123456')
            user.is_superuser = True
            user.is_staff = True
            user.is_active = True
            user.save()
            message = "admin用户密码已重置"
        
        return JsonResponse({
            'success': True,
            'message': message,
            'username': 'admin',
            'password': 'admin123456'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'重置密码失败: {str(e)}'
        })

# ========== 函数库相关API视图 ==========

@require_http_methods(["GET"])
def function_library_api(request):
    """函数库数据API"""
    try:
        # 获取所有库
        libraries = Library.objects.all()
        data = {}
        
        for library in libraries:
            library_data = {
                'name': library.library_name_cn,
                'modules': []
            }
            
            # 获取库下的所有模块
            modules = library.modules.all()
            for module in modules:
                module_data = {
                    'name': module.module_name,
                    'description': module.description or '',
                    'items': []
                }
                
                # 获取模块下的所有函数
                functions = module.functions.all()
                for function in functions:
                    function_data = {
                        'type': 'function',
                        'name': function.function_name,
                        'semantic': function.description_cn or function.description or '',
                        'operation': function.operation_type.operation_name_cn if function.operation_type else '',
                        'input': function.parameters_text or '',
                        'output': function.return_value_cn or function.return_value or ''
                    }
                    module_data['items'].append(function_data)
                
                library_data['modules'].append(module_data)
            
            data[library.library_name] = library_data
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@require_http_methods(["GET"])
def function_query_api(request):
    """函数查询API"""
    try:
        # 获取查询参数
        library_name = request.GET.get('library')
        module_name = request.GET.get('module')
        function_name = request.GET.get('function')
        operation_type = request.GET.get('operation_type')
        search = request.GET.get('search', '').lower()
        
        # 构建查询
        functions = Function.objects.select_related('module', 'module__library', 'operation_type')
        
        if library_name:
            functions = functions.filter(module__library__library_name=library_name)
        
        if module_name:
            functions = functions.filter(module__module_name=module_name)
        
        if function_name:
            functions = functions.filter(function_name__icontains=function_name)
        
        if operation_type:
            functions = functions.filter(operation_type__operation_name_cn=operation_type)
        
        if search:
            functions = functions.filter(
                models.Q(function_name__icontains=search) |
                models.Q(description__icontains=search) |
                models.Q(description_cn__icontains=search)
            )
        
        # 序列化数据
        data = []
        for function in functions:
            function_data = {
                'id': function.function_id,
                'library_name': function.module.library.library_name_cn,
                'module_name': function.module.module_name,
                'function_name': function.function_name,
                'function_name_cn': function.function_name_cn,
                'description': function.description_cn or function.description,
                'operation_type': function.operation_type.operation_name_cn if function.operation_type else '',
                'syntax': function.syntax,
                'parameters': function.parameters_text,
                'return_value': function.return_value_cn or function.return_value,
                'example': function.example_cn or function.example,
                'availability': function.availability,
                'version_added': function.version_added,
                'parameters_detail': []
            }
            
            # 获取参数详细信息
            for param in function.parameters.all():
                param_data = {
                    'name': param.parameter_name,
                    'name_cn': param.parameter_name_cn,
                    'data_type': param.data_type,
                    'is_required': param.is_required,
                    'default_value': param.default_value,
                    'description': param.description_cn or param.description
                }
                function_data['parameters_detail'].append(param_data)
            
            data.append(function_data)
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@require_http_methods(["GET"])
def function_libraries_list_api(request):
    """获取库列表API"""
    try:
        libraries = Library.objects.all()
        data = [{
            'id': library.library_id,
            'name': library.library_name,
            'name_cn': library.library_name_cn,
            'description': library.description,
            'is_builtin': library.is_builtin,
            'is_standard': library.is_standard
        } for library in libraries]
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@require_http_methods(["GET"])
def function_modules_list_api(request):
    """获取模块列表API"""
    try:
        library_id = request.GET.get('library_id')
        modules = Module.objects.select_related('library')
        
        if library_id:
            modules = modules.filter(library_id=library_id)
        
        data = [{
            'id': module.module_id,
            'name': module.module_name,
            'description': module.description,
            'library_name': module.library.library_name_cn,
            'is_builtin': module.is_builtin
        } for module in modules]
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@require_http_methods(["GET"])
def function_operation_types_api(request):
    """获取操作类型列表API"""
    try:
        operation_types = OperationType.objects.all()
        data = [{
            'id': op.operation_type_id,
            'name': op.operation_name,
            'name_cn': op.operation_name_cn,
            'description': op.description
        } for op in operation_types]
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

# ========== 函数管理相关视图 ==========

@login_required
def function_management_view(request):
    """函数管理页面视图"""
    return render(request, 'admin/函数管理.html')

@csrf_exempt
@require_http_methods(["POST"])
def upload_functions_api(request):
    """上传函数数据API"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': '没有上传文件'
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # 检查文件类型
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        file_extension = os.path.splitext(uploaded_file.name)[1].lower()
        if file_extension not in allowed_extensions:
            return JsonResponse({
                'success': False,
                'message': f'不支持的文件格式。支持格式：{", ".join(allowed_extensions)}'
            }, status=400)
        
        # 检查文件大小 (50MB = 50 * 1024 * 1024 bytes)
        max_size = 50 * 1024 * 1024
        if uploaded_file.size > max_size:
            return JsonResponse({
                'success': False,
                'message': f'文件大小超过限制。最大允许：50MB，当前文件：{uploaded_file.size / (1024*1024):.2f}MB'
            }, status=400)
        
        # 定义模板字段映射
        template_fields = {
            'Library': [
                'library_name', 'library_name_cn', 'description', 'version', 
                'category', 'is_builtin', 'is_standard'
            ],
            'Module': [
                'library_name', 'module_name', 'description', 'is_builtin'
            ],
            'Function': [
                'library_name', 'module_name', 'function_name', 'function_name_cn',
                'description', 'description_cn', 'operation_type', 'syntax',
                'parameters_text', 'return_value', 'return_value_cn', 'example',
                'example_cn', 'availability', 'version_added'
            ],
            'Parameter': [
                'library_name', 'module_name', 'function_name', 'parameter_name',
                'parameter_name_cn', 'data_type', 'is_required', 'default_value',
                'description', 'description_cn'
            ],
            'OperationType': [
                'operation_name', 'operation_name_cn', 'description'
            ],
            # 支持现有的工作表名称
            '库信息': [
                'library_name', 'library_name_cn', 'description', 'version', 
                'category', 'is_builtin', 'is_standard'
            ],
            '模块信息': [
                'library_name', 'module_name', 'description', 'is_builtin'
            ],
            '函数信息': [
                'library_name', 'module_name', 'function_name', 'function_name_cn',
                'description', 'description_cn', 'operation_type', 'syntax',
                'parameters_text', 'return_value', 'return_value_cn', 'example',
                'example_cn', 'availability', 'version_added'
            ],
            '参数信息': [
                'library_name', 'module_name', 'function_name', 'parameter_name',
                'parameter_name_cn', 'data_type', 'is_required', 'default_value',
                'description', 'description_cn'
            ],
            '操作类型': [
                'operation_name', 'operation_name_cn', 'description'
            ]
        }
        
        # 处理文件
        import pandas as pd
        import io
        
        # 读取Excel文件
        if file_extension in ['.xlsx', '.xls']:
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_names = excel_file.sheet_names
        else:
            # CSV文件处理
            content = uploaded_file.read().decode('utf-8')
            uploaded_file.seek(0)  # 重置文件指针
            df = pd.read_csv(io.StringIO(content))
            sheet_names = ['CSV_Data']
            excel_file = None
        
        processed_data = {}
        errors = []
        success_count = 0
        
        # 处理每个工作表
        for sheet_name in sheet_names:
            try:
                # 读取工作表数据
                if file_extension in ['.xlsx', '.xls']:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                else:
                    # CSV已经读取过了
                    pass
                
                # 跳过空行和说明行
                df = df.dropna(how='all')
                df = df[~df.iloc[:, 0].astype(str).str.contains('说明', na=False)]
                
                if df.empty:
                    continue
                
                # 获取表头
                headers = df.columns.tolist()
                
                # 验证字段是否匹配模板
                expected_fields = None
                for template_sheet, template_field_list in template_fields.items():
                    if sheet_name == template_sheet:
                        expected_fields = template_field_list
                        break
                
                if expected_fields is None:
                    errors.append(f"工作表 '{sheet_name}' 不在模板中，跳过处理")
                    continue
                
                # 检查字段数量
                if len(headers) != len(expected_fields):
                    errors.append(f"工作表 '{sheet_name}' 字段数量不匹配。期望：{len(expected_fields)}，实际：{len(headers)}")
                    continue
                
                # 检查字段名称（允许中文表头）
                header_mapping = {}
                for i, header in enumerate(headers):
                    # 尝试匹配字段名
                    if i < len(expected_fields):
                        header_mapping[expected_fields[i]] = header
                
                # 处理数据
                for index, row in df.iterrows():
                    try:
                        if sheet_name in ['Library', '库信息']:
                            success_count += process_library_row(row, header_mapping)
                        elif sheet_name in ['Module', '模块信息']:
                            success_count += process_module_row(row, header_mapping)
                        elif sheet_name in ['Function', '函数信息']:
                            success_count += process_function_row(row, header_mapping)
                        elif sheet_name in ['Parameter', '参数信息']:
                            success_count += process_parameter_row(row, header_mapping)
                        elif sheet_name in ['OperationType', '操作类型']:
                            success_count += process_operation_type_row(row, header_mapping)
                    except Exception as row_error:
                        errors.append(f"处理第 {index + 2} 行数据时出错：{str(row_error)}")
                
            except Exception as sheet_error:
                errors.append(f"处理工作表 '{sheet_name}' 时出错：{str(sheet_error)}")
        
        # 返回处理结果
        if errors:
            return JsonResponse({
                'success': True,
                'message': f'文件处理完成。成功导入 {success_count} 条记录。',
                'warnings': errors[:10],  # 只返回前10个错误
                'total_errors': len(errors)
            })
        else:
            return JsonResponse({
                'success': True,
                'message': f'文件处理完成。成功导入 {success_count} 条记录。'
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'上传失败: {str(e)}'
        }, status=500)

def process_library_row(row, header_mapping):
    """处理库数据行"""
    try:
        library_name = str(row[header_mapping['library_name']]).strip()
        if not library_name or library_name == 'nan':
            return 0
        
        # 检查是否已存在
        library, created = Library.objects.get_or_create(
            library_name=library_name,
            defaults={
                'library_name_cn': str(row[header_mapping['library_name_cn']]).strip() if 'library_name_cn' in header_mapping else library_name,
                'description': str(row[header_mapping['description']]).strip() if 'description' in header_mapping else '',
                'version': str(row[header_mapping['version']]).strip() if 'version' in header_mapping else '',
                'category': str(row[header_mapping['category']]).strip() if 'category' in header_mapping else '',
                'is_builtin': parse_boolean(row[header_mapping['is_builtin']]) if 'is_builtin' in header_mapping else False,
                'is_standard': parse_boolean(row[header_mapping['is_standard']]) if 'is_standard' in header_mapping else False,
            }
        )
        
        if not created:
            # 更新现有记录
            library.library_name_cn = str(row[header_mapping['library_name_cn']]).strip() if 'library_name_cn' in header_mapping else library_name
            library.description = str(row[header_mapping['description']]).strip() if 'description' in header_mapping else ''
            library.version = str(row[header_mapping['version']]).strip() if 'version' in header_mapping else ''
            library.category = str(row[header_mapping['category']]).strip() if 'category' in header_mapping else ''
            library.is_builtin = parse_boolean(row[header_mapping['is_builtin']]) if 'is_builtin' in header_mapping else False
            library.is_standard = parse_boolean(row[header_mapping['is_standard']]) if 'is_standard' in header_mapping else False
            library.save()
        
        return 1
    except Exception as e:
        raise Exception(f"处理库数据失败: {str(e)}")

def process_module_row(row, header_mapping):
    """处理模块数据行"""
    try:
        library_name = str(row[header_mapping['library_name']]).strip()
        module_name = str(row[header_mapping['module_name']]).strip()
        
        if not library_name or not module_name or library_name == 'nan' or module_name == 'nan':
            return 0
        
        # 获取库
        try:
            library = Library.objects.get(library_name=library_name)
        except Library.DoesNotExist:
            raise Exception(f"库 '{library_name}' 不存在")
        
        # 检查是否已存在
        module, created = Module.objects.get_or_create(
            library=library,
            module_name=module_name,
            defaults={
                'description': str(row[header_mapping['description']]).strip() if 'description' in header_mapping else '',
                'is_builtin': parse_boolean(row[header_mapping['is_builtin']]) if 'is_builtin' in header_mapping else False,
            }
        )
        
        if not created:
            # 更新现有记录
            module.description = str(row[header_mapping['description']]).strip() if 'description' in header_mapping else ''
            module.is_builtin = parse_boolean(row[header_mapping['is_builtin']]) if 'is_builtin' in header_mapping else False
            module.save()
        
        return 1
    except Exception as e:
        raise Exception(f"处理模块数据失败: {str(e)}")

def process_function_row(row, header_mapping):
    """处理函数数据行"""
    try:
        library_name = str(row[header_mapping['library_name']]).strip()
        module_name = str(row[header_mapping['module_name']]).strip()
        function_name = str(row[header_mapping['function_name']]).strip()
        
        if not library_name or not module_name or not function_name or 'nan' in [library_name, module_name, function_name]:
            return 0
        
        # 获取模块
        try:
            module = Module.objects.get(library__library_name=library_name, module_name=module_name)
        except Module.DoesNotExist:
            raise Exception(f"模块 '{library_name}.{module_name}' 不存在")
        
        # 获取操作类型
        operation_type = None
        if 'operation_type' in header_mapping:
            operation_type_name = str(row[header_mapping['operation_type']]).strip()
            if operation_type_name and operation_type_name != 'nan':
                operation_type, _ = OperationType.objects.get_or_create(
                    operation_name_cn=operation_type_name,
                    defaults={'operation_name': operation_type_name}
                )
        
        # 检查是否已存在
        function, created = Function.objects.get_or_create(
            module=module,
            function_name=function_name,
            defaults={
                'function_name_cn': str(row[header_mapping['function_name_cn']]).strip() if 'function_name_cn' in header_mapping else '',
                'description': str(row[header_mapping['description']]).strip() if 'description' in header_mapping else '',
                'description_cn': str(row[header_mapping['description_cn']]).strip() if 'description_cn' in header_mapping else '',
                'operation_type': operation_type,
                'syntax': str(row[header_mapping['syntax']]).strip() if 'syntax' in header_mapping else '',
                'parameters_text': str(row[header_mapping['parameters_text']]).strip() if 'parameters_text' in header_mapping else '',
                'return_value': str(row[header_mapping['return_value']]).strip() if 'return_value' in header_mapping else '',
                'return_value_cn': str(row[header_mapping['return_value_cn']]).strip() if 'return_value_cn' in header_mapping else '',
                'example': str(row[header_mapping['example']]).strip() if 'example' in header_mapping else '',
                'example_cn': str(row[header_mapping['example_cn']]).strip() if 'example_cn' in header_mapping else '',
                'availability': str(row[header_mapping['availability']]).strip() if 'availability' in header_mapping else '',
                'version_added': str(row[header_mapping['version_added']]).strip() if 'version_added' in header_mapping else '',
            }
        )
        
        if not created:
            # 更新现有记录
            function.function_name_cn = str(row[header_mapping['function_name_cn']]).strip() if 'function_name_cn' in header_mapping else ''
            function.description = str(row[header_mapping['description']]).strip() if 'description' in header_mapping else ''
            function.description_cn = str(row[header_mapping['description_cn']]).strip() if 'description_cn' in header_mapping else ''
            function.operation_type = operation_type
            function.syntax = str(row[header_mapping['syntax']]).strip() if 'syntax' in header_mapping else ''
            function.parameters_text = str(row[header_mapping['parameters_text']]).strip() if 'parameters_text' in header_mapping else ''
            function.return_value = str(row[header_mapping['return_value']]).strip() if 'return_value' in header_mapping else ''
            function.return_value_cn = str(row[header_mapping['return_value_cn']]).strip() if 'return_value_cn' in header_mapping else ''
            function.example = str(row[header_mapping['example']]).strip() if 'example' in header_mapping else ''
            function.example_cn = str(row[header_mapping['example_cn']]).strip() if 'example_cn' in header_mapping else ''
            function.availability = str(row[header_mapping['availability']]).strip() if 'availability' in header_mapping else ''
            function.version_added = str(row[header_mapping['version_added']]).strip() if 'version_added' in header_mapping else ''
            function.save()
        
        return 1
    except Exception as e:
        raise Exception(f"处理函数数据失败: {str(e)}")

def process_parameter_row(row, header_mapping):
    """处理参数数据行"""
    try:
        library_name = str(row[header_mapping['library_name']]).strip()
        module_name = str(row[header_mapping['module_name']]).strip()
        function_name = str(row[header_mapping['function_name']]).strip()
        parameter_name = str(row[header_mapping['parameter_name']]).strip()
        
        if not all([library_name, module_name, function_name, parameter_name]) or 'nan' in [library_name, module_name, function_name, parameter_name]:
            return 0
        
        # 获取函数
        try:
            function = Function.objects.get(
                module__library__library_name=library_name,
                module__module_name=module_name,
                function_name=function_name
            )
        except Function.DoesNotExist:
            raise Exception(f"函数 '{library_name}.{module_name}.{function_name}' 不存在")
        
        # 检查是否已存在
        parameter, created = Parameter.objects.get_or_create(
            function=function,
            parameter_name=parameter_name,
            defaults={
                'parameter_name_cn': str(row[header_mapping['parameter_name_cn']]).strip() if 'parameter_name_cn' in header_mapping else '',
                'data_type': str(row[header_mapping['data_type']]).strip() if 'data_type' in header_mapping else '',
                'is_required': parse_boolean(row[header_mapping['is_required']]) if 'is_required' in header_mapping else True,
                'default_value': str(row[header_mapping['default_value']]).strip() if 'default_value' in header_mapping else '',
                'description': str(row[header_mapping['description']]).strip() if 'description' in header_mapping else '',
                'description_cn': str(row[header_mapping['description_cn']]).strip() if 'description_cn' in header_mapping else '',
            }
        )
        
        if not created:
            # 更新现有记录
            parameter.parameter_name_cn = str(row[header_mapping['parameter_name_cn']]).strip() if 'parameter_name_cn' in header_mapping else ''
            parameter.data_type = str(row[header_mapping['data_type']]).strip() if 'data_type' in header_mapping else ''
            parameter.is_required = parse_boolean(row[header_mapping['is_required']]) if 'is_required' in header_mapping else True
            parameter.default_value = str(row[header_mapping['default_value']]).strip() if 'default_value' in header_mapping else ''
            parameter.description = str(row[header_mapping['description']]).strip() if 'description' in header_mapping else ''
            parameter.description_cn = str(row[header_mapping['description_cn']]).strip() if 'description_cn' in header_mapping else ''
            parameter.save()
        
        return 1
    except Exception as e:
        raise Exception(f"处理参数数据失败: {str(e)}")

def process_operation_type_row(row, header_mapping):
    """处理操作类型数据行"""
    try:
        operation_name = str(row[header_mapping['operation_name']]).strip()
        if not operation_name or operation_name == 'nan':
            return 0
        
        # 检查是否已存在
        operation_type, created = OperationType.objects.get_or_create(
            operation_name=operation_name,
            defaults={
                'operation_name_cn': str(row[header_mapping['operation_name_cn']]).strip() if 'operation_name_cn' in header_mapping else operation_name,
                'description': str(row[header_mapping['description']]).strip() if 'description' in header_mapping else '',
            }
        )
        
        if not created:
            # 更新现有记录
            operation_type.operation_name_cn = str(row[header_mapping['operation_name_cn']]).strip() if 'operation_name_cn' in header_mapping else operation_name
            operation_type.description = str(row[header_mapping['description']]).strip() if 'description' in header_mapping else ''
            operation_type.save()
        
        return 1
    except Exception as e:
        raise Exception(f"处理操作类型数据失败: {str(e)}")

def parse_boolean(value):
    """解析布尔值"""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        value = value.strip().lower()
        if value in ['true', '1', 'yes', '是', 't']:
            return True
        elif value in ['false', '0', 'no', '否', 'f']:
            return False
    return False

@require_http_methods(["GET"])
def get_function_stats_api(request):
    """获取函数统计信息API"""
    try:
        from django.db.models import Count
        
        stats = {
            'libraries': Library.objects.count(),
            'modules': Module.objects.count(),
            'functions': Function.objects.count(),
            'parameters': Parameter.objects.count()
        }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'获取统计信息失败: {str(e)}'
        }, status=500)

@require_http_methods(["GET"])
def get_table_structure_api(request):
    """获取数据表结构API"""
    try:
        from django.db import connection
        
        # 定义表结构信息
        table_structures = {
            'function_library': {
                'name': '函数库 (Library)',
                'fields': [
                    'library_id (AutoField)',
                    'library_name (CharField)',
                    'library_name_cn (CharField)',
                    'description (TextField)',
                    'version (CharField)',
                    'category (CharField)',
                    'is_builtin (BooleanField)',
                    'is_standard (BooleanField)',
                    'created_at (DateTimeField)',
                    'updated_at (DateTimeField)'
                ]
            },
            'function_module': {
                'name': '模块 (Module)',
                'fields': [
                    'module_id (AutoField)',
                    'library_id (ForeignKey)',
                    'module_name (CharField)',
                    'description (TextField)',
                    'is_builtin (BooleanField)',
                    'created_at (DateTimeField)',
                    'updated_at (DateTimeField)'
                ]
            },
            'function_info': {
                'name': '函数 (Function)',
                'fields': [
                    'function_id (AutoField)',
                    'module_id (ForeignKey)',
                    'function_name (CharField)',
                    'function_name_cn (CharField)',
                    'description (TextField)',
                    'description_cn (TextField)',
                    'operation_type_id (ForeignKey)',
                    'syntax (TextField)',
                    'parameters_text (TextField)',
                    'return_value (TextField)',
                    'return_value_cn (TextField)',
                    'example (TextField)',
                    'example_cn (TextField)',
                    'availability (CharField)',
                    'version_added (CharField)',
                    'created_at (DateTimeField)',
                    'updated_at (DateTimeField)'
                ]
            },
            'function_parameter': {
                'name': '参数 (Parameter)',
                'fields': [
                    'parameter_id (AutoField)',
                    'function_id (ForeignKey)',
                    'parameter_name (CharField)',
                    'parameter_name_cn (CharField)',
                    'data_type (CharField)',
                    'is_required (BooleanField)',
                    'default_value (CharField)',
                    'description (TextField)',
                    'description_cn (TextField)'
                ]
            },
            'function_operation_type': {
                'name': '操作类型 (OperationType)',
                'fields': [
                    'operation_type_id (AutoField)',
                    'operation_name (CharField)',
                    'operation_name_cn (CharField)',
                    'description (TextField)'
                ]
            }
        }
        
        return JsonResponse({
            'success': True,
            'tables': table_structures
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'获取表结构失败: {str(e)}'
        }, status=500)

@require_http_methods(["GET"])
def export_functions_api(request):
    """导出函数数据API"""
    try:
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        # 获取查询参数
        format_type = request.GET.get('format', 'xlsx')  # 默认xlsx格式
        
        if format_type == 'csv':
            # CSV格式导出
            import csv
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="函数数据库表.csv"'
            response.write('\ufeff')  # 添加BOM以支持中文
            
            writer = csv.writer(response)
            writer.writerow(['库名称', '模块名称', '函数名称', '函数中文名称', '描述', '操作类型', '语法', '参数', '返回值', '示例'])
            
            functions = Function.objects.select_related('module', 'module__library', 'operation_type')
            
            for function in functions:
                writer.writerow([
                    function.module.library.library_name_cn,
                    function.module.module_name,
                    function.function_name,
                    function.function_name_cn or '',
                    function.description_cn or function.description or '',
                    function.operation_type.operation_name_cn if function.operation_type else '',
                    function.syntax or '',
                    function.parameters_text or '',
                    function.return_value_cn or function.return_value or '',
                    function.example_cn or function.example or ''
                ])
            
            return response
        
        else:
            # XLSX格式导出
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "函数数据库表"
            
            # 定义表头
            headers = [
                '库名称', '模块名称', '函数名称', '函数中文名称', '描述', 
                '操作类型', '语法', '参数', '返回值', '示例'
            ]
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 获取数据
            functions = Function.objects.select_related('module', 'module__library', 'operation_type')
            
            # 写入数据
            for row, function in enumerate(functions, 2):
                ws.cell(row=row, column=1, value=function.module.library.library_name_cn)
                ws.cell(row=row, column=2, value=function.module.module_name)
                ws.cell(row=row, column=3, value=function.function_name)
                ws.cell(row=row, column=4, value=function.function_name_cn or '')
                ws.cell(row=row, column=5, value=function.description_cn or function.description or '')
                ws.cell(row=row, column=6, value=function.operation_type.operation_name_cn if function.operation_type else '')
                ws.cell(row=row, column=7, value=function.syntax or '')
                ws.cell(row=row, column=8, value=function.parameters_text or '')
                ws.cell(row=row, column=9, value=function.return_value_cn or function.return_value or '')
                ws.cell(row=row, column=10, value=function.example_cn or function.example or '')
            
            # 自动调整列宽
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                column = ws[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 设置数据行样式
            data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for row in range(2, ws.max_row + 1):
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = data_alignment
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 创建响应
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="函数数据库表.xlsx"'
            
            return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'导出失败: {str(e)}'
        }, status=500)

@require_http_methods(["GET"])
def download_template_api(request):
    """下载数据模板API"""
    try:
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 创建5个工作表，对应5个数据表
        tables = [
            {
                'name': '函数库表',
                'sheet_name': 'Library',
                'headers': [
                    'library_name', 'library_name_cn', 'description', 'version', 
                    'category', 'is_builtin', 'is_standard'
                ],
                'header_names': [
                    '库名称(英文)', '库名称(中文)', '描述', '版本', 
                    '分类', '是否内置', '是否标准库'
                ],
                'sample_data': [
                    ['os', '操作系统接口', '提供操作系统相关功能', '3.x', '系统', True, True],
                    ['sys', '系统相关参数', '系统相关的参数和函数', '3.x', '系统', True, True]
                ]
            },
            {
                'name': '模块表',
                'sheet_name': 'Module',
                'headers': [
                    'library_name', 'module_name', 'description', 'is_builtin'
                ],
                'header_names': [
                    '所属库名称', '模块名称', '描述', '是否内置'
                ],
                'sample_data': [
                    ['os', 'path', '路径操作相关函数', True],
                    ['os', 'file', '文件操作相关函数', True]
                ]
            },
            {
                'name': '函数表',
                'sheet_name': 'Function',
                'headers': [
                    'library_name', 'module_name', 'function_name', 'function_name_cn',
                    'description', 'description_cn', 'operation_type', 'syntax',
                    'parameters_text', 'return_value', 'return_value_cn', 'example',
                    'example_cn', 'availability', 'version_added'
                ],
                'header_names': [
                    '所属库名称', '所属模块名称', '函数名称(英文)', '函数名称(中文)',
                    '描述(英文)', '描述(中文)', '操作类型', '语法',
                    '参数文本', '返回值(英文)', '返回值(中文)', '示例(英文)',
                    '示例(中文)', '可用性', '版本'
                ],
                'sample_data': [
                    ['os', 'path', 'join', '路径拼接', 'Join path components', '拼接路径组件', '路径操作', 'os.path.join(path, *paths)', 'path, *paths', 'str', '字符串', 'os.path.join("a", "b")', 'os.path.join("a", "b")', 'All', '1.4']
                ]
            },
            {
                'name': '参数表',
                'sheet_name': 'Parameter',
                'headers': [
                    'library_name', 'module_name', 'function_name', 'parameter_name',
                    'parameter_name_cn', 'data_type', 'is_required', 'default_value',
                    'description', 'description_cn'
                ],
                'header_names': [
                    '所属库名称', '所属模块名称', '所属函数名称', '参数名称(英文)',
                    '参数名称(中文)', '数据类型', '是否必需', '默认值',
                    '描述(英文)', '描述(中文)'
                ],
                'sample_data': [
                    ['os', 'path', 'join', 'path', '路径', 'str', True, '', 'Base path', '基础路径'],
                    ['os', 'path', 'join', '*paths', '路径列表', 'str', False, '', 'Additional paths', '额外路径']
                ]
            },
            {
                'name': '操作类型表',
                'sheet_name': 'OperationType',
                'headers': [
                    'operation_name', 'operation_name_cn', 'description'
                ],
                'header_names': [
                    '操作类型(英文)', '操作类型(中文)', '描述'
                ],
                'sample_data': [
                    ['File Operation', '文件操作', '文件相关的操作类型'],
                    ['Path Operation', '路径操作', '路径相关的操作类型']
                ]
            }
        ]
        
        # 为每个表创建工作表
        for table in tables:
            # 删除默认工作表（如果是第一个表）
            if table == tables[0]:
                ws = wb.active
                ws.title = table['sheet_name']
            else:
                ws = wb.create_sheet(title=table['sheet_name'])
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(table['header_names'], 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 写入示例数据
            for row, data in enumerate(table['sample_data'], 2):
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 添加说明行
            note_row = len(table['sample_data']) + 3
            ws.cell(row=note_row, column=1, value="说明：")
            ws.cell(row=note_row, column=1).font = Font(bold=True, color="FF0000")
            
            note_row += 1
            ws.cell(row=note_row, column=1, value="1. 请按照示例格式填写数据")
            note_row += 1
            ws.cell(row=note_row, column=1, value="2. 布尔值请填写：True 或 False")
            note_row += 1
            ws.cell(row=note_row, column=1, value="3. 必填字段不能为空")
            note_row += 1
            ws.cell(row=note_row, column=1, value="4. 删除示例数据行后再上传")
            
            # 自动调整列宽
            for col in range(1, len(table['headers']) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                column = ws[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 40)  # 最大宽度40
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 删除默认工作表（如果还存在）
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 创建响应
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="函数数据库模板.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'下载模板失败: {str(e)}'
        }, status=500)